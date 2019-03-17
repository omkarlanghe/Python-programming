import openpyxl
import sys
path = r"/home/omkar/Desktop/names_input.xlsx"
vcf=open(r"/home/omkar/Desktop/hello.vcf","w")
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 1, column = 1)

for j in range(1,6):
	for i in range(1, 5): 
		cell_obj = sheet_obj.cell(row = j, column = i) 
		print(cell_obj.value),
 	print("")
 
for j in range(1,6):
	name= sheet_obj.cell(row = j, column = 1).value
	email= sheet_obj.cell(row = j, column = 2).value
	contact= sheet_obj.cell(row = j, column = 3).value
	collage= sheet_obj.cell(row = j, column = 4).value	
 	vcf.write("BEGIN:VCARD"+"\n")
 	vcf.write("VESION:3.0"+"\n")
 	
 	vcf.write("FN: "+name+"\n")
 	vcf.write("ORG: "+collage+"\n")
 	vcf.write("TEL: "+str(contact)+"\n")
 	vcf.write("EMAIL: "+email+"\n")
 	vcf.write("END:VCARD"+"\n")	
 	
vcf.close()
wb_obj.close()	


